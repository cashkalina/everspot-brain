"""Golden-file test (SPEC §12) — synthetic ``acme_synth`` flat register.

Re-runs the deterministic spine (ingest → assemble → emit → dry-load) on the
SYNTHETIC fixture and diffs the produced artifacts against the committed golden.
A future change that alters output fails LOUDLY with a diff — this is the safety net
that gates self-modification (§11.4).

The fixture deliberately exercises:
  - burial split: one register row → Property + Customer + Interment;
  - multi-occupancy parent dedup: rows sharing PLOT_NO A-12-1 → ONE property, TWO interments;
  - partial dates: a 0-placeholder month/day → benign null; Feb-29-1990 (non-leap) →
    day dropped to null + a data_quality flag;
  - a value-set resolution: ITYPE BUR/CRE → interment_type_id 11/12;
  - an empty grave (A-13-1) → property-only, no phantom child records.
"""

import json
import sys
from pathlib import Path

import pytest

GOLDEN = Path(__file__).resolve().parent
sys.path.insert(0, str(GOLDEN))  # make tests/golden/spine.py importable

from spine import run_spine  # noqa: E402

EXPECTED = GOLDEN / "expected" / "acme_synth"
ENTITIES = ["property_group", "property", "customer", "interment"]


@pytest.fixture(scope="module")
def spine(tmp_path_factory):
    dest = tmp_path_factory.mktemp("golden")
    return run_spine("acme_synth", dest)


def _read_ndjson_sorted(path: Path) -> list[dict]:
    rows = [json.loads(line) for line in path.read_text().splitlines() if line.strip()]
    rows.sort(key=lambda r: r["external_id"])
    return rows


@pytest.mark.parametrize("entity", ENTITIES)
def test_canonical_ndjson_matches_golden(spine, entity):
    produced = _read_ndjson_sorted(spine["canonical_dir"] / f"{entity}.ndjson")
    expected = _read_ndjson_sorted(EXPECTED / "canonical" / f"{entity}.ndjson")
    assert produced == expected, f"canonical {entity}.ndjson drifted from golden"


def test_entity_counts(spine):
    counts = spine["assemble_result"].entity_counts
    assert counts == {"property_group": 1, "property": 3, "customer": 3, "interment": 3}


def test_multi_occupancy_property_dedup(spine):
    """Two register rows sharing a plot → ONE property, TWO interments linked to it.

    external_ids are opaque hashes, so identify the shared property by deduped
    property_ref (every interment under one plot points at the same external_id)
    rather than by parsing the plot out of the id.
    """
    interments = _read_ndjson_sorted(spine["canonical_dir"] / "interment.ndjson")
    by_ref: dict[str, int] = {}
    for i in interments:
        by_ref[i["property_ref"]] = by_ref.get(i["property_ref"], 0) + 1
    shared_ref = max(by_ref, key=by_ref.get)
    assert by_ref[shared_ref] == 2  # two interments share one property
    properties = _read_ndjson_sorted(spine["canonical_dir"] / "property.ndjson")
    a12_1 = [p for p in properties if p["external_id"] == shared_ref]
    assert len(a12_1) == 1  # deduped to a single property


def test_feb_29_non_leap_day_dropped_and_flagged(spine):
    interments = _read_ndjson_sorted(spine["canonical_dir"] / "interment.ndjson")
    # opaque external_ids — locate Mary by her (internal, readable) source_id.
    mary = next(i for i in interments if i["_provenance"]["source_id"] == "register:A-12-1|2")
    assert mary["dod"] == {"year": 1990, "month": 2, "day": None, "estimated": True}
    flags = [na.to_dict() for na in spine["assemble_result"].needs_attention]
    assert any(f["kind"] == "data_quality" and "dod" in f["detail"] for f in flags)


def test_empty_grave_is_property_only_no_phantom_children(spine):
    customers = _read_ndjson_sorted(spine["canonical_dir"] / "customer.ndjson")
    interments = _read_ndjson_sorted(spine["canonical_dir"] / "interment.ndjson")
    # A-13-1 produced a property but NO customer/interment. external_ids are opaque,
    # so check the (internal, readable) source_id provenance, not the id substring.
    assert not any("A-13-1" in c["_provenance"]["source_id"] for c in customers)
    assert not any("A-13-1" in i["_provenance"]["source_id"] for i in interments)


def test_value_set_resolution(spine):
    interments = _read_ndjson_sorted(spine["canonical_dir"] / "interment.ndjson")
    by_src = {i["_provenance"]["source_id"]: i for i in interments}
    assert by_src["register:A-12-1|1"]["interment_type_id"] == 11  # BUR
    assert by_src["register:A-12-2|3"]["interment_type_id"] == 12  # CRE


def test_emit_produces_expected_wave_files_and_columns(spine):
    from openpyxl import load_workbook

    expected = json.loads((EXPECTED / "emit_summary.json").read_text())
    produced = {}
    for p in spine["emit_files"]:
        wb = load_workbook(p)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        produced[p.name] = {"header": list(rows[0]), "data_rows": len(rows) - 1}
    assert produced == expected


def test_dry_load_plan_matches_golden(spine):
    expected = json.loads((EXPECTED / "dry_load_plan.json").read_text())
    dc = spine["dry_client"]
    res = spine["load_result"]
    produced = {
        "created_counts": res.created,
        "updated_counts": res.updated,
        "failed_counts": res.failed,
        "registered_external_ids": sorted(e for (_m, e) in dc.registered),
        "errors": res.errors,
    }
    assert produced == expected


def test_dry_load_registers_no_duplicate_external_ids(spine):
    regs = [e for (_m, e) in spine["dry_client"].registered]
    assert len(regs) == len(set(regs))
